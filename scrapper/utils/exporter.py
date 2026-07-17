import os
import json
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config
from utils.logger import get_logger

logger = get_logger("exporter")

COLUMNS = [
    "website",
    "product_name",
    "brand",
    "price",
    "old_price",
    "discount",
    "availability",
    "product_url",
    "product_image",
]


def _build_dataframe(results: list) -> pd.DataFrame:
    if not results:
        return pd.DataFrame(columns=COLUMNS)
    df = pd.DataFrame(results)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNS]


def export_csv(results: list, filepath: str) -> None:
    df = _build_dataframe(results)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    logger.info(f"CSV exported to {filepath}")


def export_json(results: list, filepath: str) -> None:
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON exported to {filepath}")


def export_excel(results: list, filepath: str) -> None:
    df = _build_dataframe(results)
    df.to_excel(filepath, index=False, sheet_name="Results", engine="openpyxl")

    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    ws = wb["Results"]

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not df.empty:
        cheapest_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cheapest_font = Font(bold=True, color="006100")
        min_price_row = df["price"].astype(float).idxmin() + 2
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=min_price_row, column=col_idx)
            cell.fill = cheapest_fill
            cell.font = cheapest_font

    for col_idx, column in enumerate(COLUMNS, start=1):
        max_len = max([len(str(column))] + [len(str(v)) for v in df[column].tolist()]) if not df.empty else len(column)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    logger.info(f"Excel exported to {filepath}")


def export_all(results: list, base_name: str = "results") -> dict:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"{base_name}_{timestamp}"

    paths = {
        "csv": os.path.join(config.OUTPUT_DIR, f"{prefix}.csv"),
        "json": os.path.join(config.OUTPUT_DIR, f"{prefix}.json"),
        "xlsx": os.path.join(config.OUTPUT_DIR, f"{prefix}.xlsx"),
    }

    try:
        export_csv(results, paths["csv"])
    except Exception as exc:
        logger.error(f"Failed to export CSV: {exc}")

    try:
        export_json(results, paths["json"])
    except Exception as exc:
        logger.error(f"Failed to export JSON: {exc}")

    try:
        export_excel(results, paths["xlsx"])
    except Exception as exc:
        logger.error(f"Failed to export Excel: {exc}")

    return paths
